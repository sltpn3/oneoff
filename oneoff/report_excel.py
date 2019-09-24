data = [{"n_influencer": ["ali+jamil", "nurcahyo+nugroho"], "n_id": "2539861131153409", "m_displayName": "Serikat News", "n_date": "2019-03-10 23:18:18", "n_sent": 1, "n_link": "http://serikatnews.com/kayu-lapis-asal-banten-tembus-pasar-dunia/", "n_title": "Kayu Lapis Asal Banten Tembus Pasar Dunia"}, {"n_influencer": ["mat+sahri"], "n_id": "2539857494691840", "m_displayName": "Trubus.id", "n_date": "2019-03-10 23:03:37", "n_sent": 0, "n_link": "https://news.trubus.id/baca/26829/kementan-bagi-bagi-instalasi-pengering-uv-ke-petani-di-sumenep", "n_title": "Kementan Bagi\u2013bagi Instalasi Pengering UV ke Petani di Sumenep"}, {"n_influencer": ["khofifah+indar+parawansa", "artomoro+rianto"], "n_id": "2539857473720321", "m_displayName": "Harian Bhirawa", "n_date": "2019-03-10 23:03:08", "n_sent": 0, "n_link": "http://harianbhirawa.com/pemprov-jatim-siap-membuat-regulasi-larangan-ekspor-biji-porang/", "n_title": "Pemprov Jatim Siap Membuat Regulasi Larangan Ekspor Biji Porang"}, {"n_influencer": ["suzi+marsitawati", "rahmat+zai"], "n_id": "2539850892857346", "m_displayName": "Antara", "n_date": "2019-03-10 22:27:01", "n_sent": 1, "n_link": "https://www.antaranews.com/berita/807870/mengakhiri-eksploitasi-kera-ekor-panjang", "n_title": "Mengakhiri eksploitasi kera ekor panjang"}, {"n_influencer": ["basmin+mattayang"], "n_id": "2539847151538179", "m_displayName": "Kabar News", "n_date": "2019-03-10 22:22:27", "n_sent": 1, "n_link": "http://kabar.news/besok-mentan-ke-luwu-pemkab-telah-rampungkan-segala-persiapan", "n_title": "Besok, Mentan ke Luwu, Pemkab telah Rampungkan Segala Persiapan"}, {"n_influencer": ["darmin+nasution", "joko+widodo", "herman+deru"], "n_id": "2539846098767874", "m_displayName": "Bisnis Indonesia", "n_date": "2019-03-10 22:20:07", "n_sent": 0, "n_link": "https://ekonomi.bisnis.com/read/20190310/9/897987/penggunaan-karet-untuk-aspal-bakal-diwajibkan-ke-seluruh-daerah", "n_title": "Penggunaan Karet untuk Aspal Bakal Diwajibkan ke Seluruh Daerah"}, {"n_influencer": ["khofifah+indar+parawansa"], "n_id": "2539846061019137", "m_displayName": "RRI News", "n_date": "2019-03-10 22:17:36", "n_sent": 1, "n_link": "http://rri.co.id/post/berita/646666/daerah/pemprov_jatim_perjuangkan_legalitas_155_lmdh.html", "n_title": "Pemprov Jatim Perjuangkan Legalitas 155 LMDH"}, {"n_influencer": ["ahmad+munawir"], "n_id": "2539843741569026", "m_displayName": "Antara Jabar", "n_date": "2019-03-10 22:02:47", "n_sent": 1, "n_link": "https://jabar.antaranews.com/nasional/berita/807844/menghalau-topeng-monyet?utm_source=antaranews&utm_medium=nasional&utm_campaign=antaranews", "n_title": "Menghalau topeng monyet"}, {"n_influencer": ["ahmad+munawir"], "n_id": "2539839790534657", "m_displayName": "Antara", "n_date": "2019-03-10 21:43:16", "n_sent": 1, "n_link": "https://www.antaranews.com/berita/807844/menghalau-topeng-monyet", "n_title": "Menghalau topeng monyet"}, {"n_influencer": ["syamsul+s+p", "laode+masihu+kamaluddin", "suprapti"], "n_id": "2539835906609152", "m_displayName": "Detik Sultra", "n_date": "2019-03-10 21:38:44", "n_sent": 0, "n_link": "https://detiksultra.com/pemerintah-bangun-lima-gudang-alsintan-di-indonesia-salah-satunya-di-konsel", "n_title": "Pemerintah Bangun Lima Gudang Alsintan di Indonesia, Salah Satunya di Konsel"}]

from openpyxl import Workbook, load_workbook

wb = load_workbook('template_online_news_A.xlsx')
ws = wb.active
# ws['A2'] = '123'



row = 2
num = 1

for news in data:
    ws['A{}'.format(row)] = num
    ws['B{}'.format(row)] = news['n_date']
    ws['C{}'.format(row)] = news['n_title']
    ws['D{}'.format(row)] = news['m_displayName']
    ws['E{}'.format(row)] = news['n_sent']
    ws['F{}'.format(row)] = ', '.join(news['n_influencer'])
    ws['G{}'.format(row)] = news['n_link']
    row += 1
    num += 1
    
wb.save('test.xlsx')